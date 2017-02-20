from import_data_with_color import importdata
from network_framework_auto import framework
from Record_result import record_label, record_error
import numpy as np
import tensorflow as tf
import openpyxl


def nn_test(test_option,framework_file):
    # Set model path
    model_load_name = test_option['model_load_name']
    model_load_path = './' + test_option['model_path'] + '/' + model_load_name

    # Set parameters
    data_name = test_option['data_name']
    num_class = test_option['num_class']
    test_batch_size = test_option['batch_size']
    keep_prob = test_option['keep_prob']
    save_label = test_option['save_record']
    record_file_name = test_option['record_name']

    # Import test data
    (feature_test, label_test) = importdata(data_name)

    # Build up network framework
    model = framework(feature_test.shape, num_class, framework_data=framework_file, keep_prob=keep_prob)

    # Initialize saver
    saver = tf.train.Saver()

    # Create optimizer and session
    sess = tf.InteractiveSession()
    tf.global_variables_initializer().run()

    # Load per-trained model
    saver.restore(sess, model_load_path)
    print('Load model =', model_load_path)

    # Initialize records
    if save_label:
        record_labels = np.zeros((label_test.shape[1],label_test.shape[1]))

    # Test
    correct_prediction = tf.equal(tf.argmax(model.y,1),tf.argmax(model.y_pred,1))
    correct_batch = tf.reduce_sum(tf.cast(correct_prediction,tf.float32))
    correct_total = 0.

    # Record errors
    error_list = [['Number', 'Predicted', 'True']]

    for i in range(int(feature_test.shape[0] / test_batch_size) + 1):
        batch_feature = feature_test[test_batch_size * i:min(test_batch_size * (i + 1) - 1, feature_test.shape[0])]
        batch_label = label_test[test_batch_size * i:min(test_batch_size * (i + 1) - 1, label_test.shape[0])]
        correct_batch_get, batch_pred_label = sess.run([correct_batch, model.y_pred], feed_dict={model.x: batch_feature, model.y: batch_label})
        correct_total += correct_batch_get

        # Save label output times
        if save_label:
            record_labels = record_labels + record_label(batch_pred_label, batch_label)
            error_tem = record_error(batch_pred_label, batch_label, i * test_batch_size)
            if not error_tem == []:
                error_list = np.append(error_list, error_tem, axis=0)

    # Calculate accuracy
    accuracy = correct_total / feature_test.shape[0]
    print(error_list.shape)

    # Save model
    if save_label:
        # Open work file
        workbook = openpyxl.load_workbook(filename=record_file_name)

        # Create work sheet
        worksheet = workbook.create_sheet(title='test')

        # Record
        worksheet['A1'] = 'Test data'
        worksheet['B1'] = data_name
        worksheet['A2'] = 'Accuracy'
        worksheet['B2'] = '%04f' % accuracy
        worksheet['A3'] = 'Use model'
        worksheet['B3'] = model_load_name

        # Create work sheet
        worksheet = workbook.create_sheet(title='labels')

        # Record labels
        for i in range(record_labels.shape[0]):
            worksheet.cell(row=i + 2, column=1, value=i + 1)
            worksheet.cell(row=1, column=i + 2, value=i + 1)
            for j in range(record_labels.shape[1]):
                worksheet.cell(row=i + 2, column=j + 2, value=record_labels[i,j])

        # Create work sheet
        worksheet = workbook.create_sheet(title='errors')

        for i in range(error_list.shape[0]):
            worksheet.cell(row=i + 1, column=1, value=error_list[i, 0])
            worksheet.cell(row=i + 1, column=2, value=error_list[i, 1])
            worksheet.cell(row=i + 1, column=3, value=error_list[i, 2])

        # Save record
        workbook.save(filename=record_file_name)

    # Display accuracy
    print("Accuracy = ", accuracy)

    return accuracy